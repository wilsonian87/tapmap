"""XLSX export with proper formatting for measurement teams.

Produces a spreadsheet with:
- Frozen header row with auto-filter
- Column widths tuned for content
- Blank MVA/HVA scoring columns for manual annotation
- Conditional formatting for pharma context hints
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Column definitions: (header, db_field, width)
COLUMNS = [
    ("Page URL", "page_url", 45),
    ("Page Title", "page_title", 30),
    ("Element Type", "element_type", 14),
    ("Action Type", "action_type", 14),
    ("Element Text", "element_text", 40),
    ("CSS Selector", "css_selector", 35),
    ("Section Context", "section_context", 30),
    ("Container", "container_context", 12),
    ("Above Fold", "is_above_fold", 11),
    ("Target URL", "target_url", 40),
    ("External", "is_external", 10),
    ("Pharma Context", "pharma_context", 18),
    # Blank scoring columns for manual annotation
    ("Value Tier", "value_tier", 12),
    ("Value Reason", "value_reason", 25),
    ("Owner", "owner", 15),
    ("Measurement Status", "measurement_status", 18),
]

HEADER_FONT = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="6B21A8", end_color="6B21A8", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Calibri", size=10)
DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=False)
WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

PHARMA_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
EXTERNAL_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

THIN_BORDER = Border(
    bottom=Side(style="thin", color="E5E7EB"),
)


def generate_xlsx(
    elements: list[dict],
    scan_info: dict,
) -> io.BytesIO:
    """Generate an XLSX workbook from extracted elements.

    Returns a BytesIO buffer containing the workbook.
    """
    wb = Workbook()

    # === Elements Sheet ===
    ws = wb.active
    ws.title = "Elements"

    # Write headers
    for col_idx, (header, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data rows
    for row_idx, element in enumerate(elements, 2):
        for col_idx, (_, field, _) in enumerate(COLUMNS, 1):
            value = element.get(field)

            # Convert booleans to readable text
            if field == "is_above_fold":
                value = "Yes" if value else "No"
            elif field == "is_external":
                value = "Yes" if value else "No"

            cell = ws.cell(row=row_idx, column=col_idx, value=value or "")
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

            # Wrap text for long content columns
            if field in ("element_text", "css_selector", "page_url", "target_url"):
                cell.alignment = WRAP_ALIGNMENT
            else:
                cell.alignment = DATA_ALIGNMENT

        # Highlight pharma context rows
        pharma_ctx = element.get("pharma_context")
        if pharma_ctx:
            pharma_col = next(
                i for i, (_, f, _) in enumerate(COLUMNS, 1) if f == "pharma_context"
            )
            ws.cell(row=row_idx, column=pharma_col).fill = PHARMA_FILL

        # Highlight external links
        if element.get("is_external"):
            ext_col = next(
                i for i, (_, f, _) in enumerate(COLUMNS, 1) if f == "is_external"
            )
            ws.cell(row=row_idx, column=ext_col).fill = EXTERNAL_FILL

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    if elements:
        last_col = get_column_letter(len(COLUMNS))
        ws.auto_filter.ref = f"A1:{last_col}{len(elements) + 1}"

    # Set row height for header
    ws.row_dimensions[1].height = 30

    # === Summary Sheet ===
    ws_summary = wb.create_sheet("Scan Summary")

    summary_data = [
        ("Scan ID", scan_info.get("scan_id", "")),
        ("Domain", scan_info.get("domain", "")),
        ("URL", scan_info.get("scan_url", "")),
        ("Date", scan_info.get("crawl_date", "")),
        ("Status", scan_info.get("scan_status", "")),
        ("Pages Scanned", scan_info.get("pages_scanned", 0)),
        ("Total Elements", len(elements)),
        ("Duration (sec)", scan_info.get("duration_seconds", "")),
        ("Scan Quality", scan_info.get("scan_quality", "")),
        ("Consent Detected", "Yes" if scan_info.get("consent_detected") else "No"),
        ("Consent Action", scan_info.get("consent_action", "")),
        ("Consent Framework", scan_info.get("consent_framework", "")),
        ("robots.txt Found", "Yes" if scan_info.get("robots_txt_found") else "No"),
        ("robots.txt Respected", "Yes" if scan_info.get("robots_txt_respected") else "No"),
    ]

    # Type breakdown
    type_counts = {}
    pharma_count = 0
    for el in elements:
        t = el.get("element_type", "unknown")
        type_counts[t] = type_counts.get(t, 0) + 1
        if el.get("pharma_context"):
            pharma_count += 1

    summary_data.append(("", ""))
    summary_data.append(("Element Breakdown", ""))
    for t, count in sorted(type_counts.items()):
        summary_data.append((f"  {t}", count))
    summary_data.append(("  pharma-flagged", pharma_count))

    for row_idx, (label, value) in enumerate(summary_data, 1):
        label_cell = ws_summary.cell(row=row_idx, column=1, value=label)
        value_cell = ws_summary.cell(row=row_idx, column=2, value=value)
        label_cell.font = Font(name="Calibri", bold=True, size=10)
        value_cell.font = Font(name="Calibri", size=10)

    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 50

    # Write to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
